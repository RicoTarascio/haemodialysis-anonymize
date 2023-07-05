from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
import os
from collections.abc import Iterator
from uuid import uuid4, NAMESPACE_DNS
from dataclasses import dataclass
from shutil import copy2
import datetime


@dataclass
class Patient:
    name: str
    surname: str
    birthDate: str


ROOT_DIR = os.path.abspath(os.curdir)
ORIGINAL_FOLDER_PATH = os.path.join(ROOT_DIR, "tests/original")

ANON_FOLDER_PATH = os.path.join(ROOT_DIR, "tests/anonymized")
ANON_ENTRIES_PATH = os.path.join(ANON_FOLDER_PATH, "anon_entries.xlsx")


def get_anon_entries_wb():
    if not os.path.exists(ANON_ENTRIES_PATH):
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Entries"
            wss = wb.get_sheet_by_name("Entries")
            wss.append(["Nome", "Cognome", "Data di nascita", "ID"])
            for c in ["A1", "B1", "C1", "D1"]:
                wss[c].font = Font(size=20, bold=True)
                wss[c].alignment = Alignment(horizontal="center")
            wb.save(ANON_ENTRIES_PATH)
            return wb
    else:
        return load_workbook(ANON_ENTRIES_PATH)


def get_dir(path: str):
    return list(os.scandir(path))


def get_patients(dir: list[os.DirEntry]) -> list[Patient]:
    # For each file: name-surname-DD_MM__YYYY
    patients = []
    for f in dir:
        noExt = str(f.name).split(".")[0]
        [name, surname, birthDate] = noExt.strip().upper().split("-")
        patients.append(Patient(name, surname, birthDate))
    return patients


def generate_UUID4():
    return "".join(str(uuid4()).split("-"))[:10]


def save_anon_patient_file(file: os.DirEntry, patientID: str):
    anonFileName = "".join(
        [patientID, "_", str(datetime.datetime.today().year), ".xlsx"]
    )
    destinationPath = os.path.join(ANON_FOLDER_PATH, anonFileName)
    copy2(file.path, destinationPath)


def is_patient_in_ws(ws: Worksheet, patient: Patient):
    for i, row in enumerate(ws.values):
        if (
            row[0] == patient.name
            and row[1] == patient.surname
            and row[2] == patient.birthDate
        ):
            print("Patient: " + str(patient) + " found existing in row " + str(i))
            return True
    return False


def get_patient_id(ws: Worksheet, patient: Patient):
    for i, row in enumerate(ws.values):
        if (
            row[0] == patient.name
            and row[1] == patient.surname
            and row[2] == patient.birthDate
        ):
            print(
                "Found id: "
                + str(row[3])
                + " for patient "
                + str(patient)
                + " in row: "
                + str(i)
            )
            return str(row[3])
    return None


def create_anon_entry(wb: Workbook, ws: Worksheet, patient: Patient, ID: str):
    ws.append([patient.name, patient.surname, patient.birthDate, ID])


def anonymize():
    os.makedirs(ANON_FOLDER_PATH, exist_ok=True)
    if not os.path.exists(ORIGINAL_FOLDER_PATH):
        print(
            "[WARNING] Original files folder was not found, one will be created for you in tests/original. This execution will be exited"
        )
        os.makedirs(ORIGINAL_FOLDER_PATH, exist_ok=True)
        return
    anonWb = get_anon_entries_wb()
    anonWs = None
    if anonWb is not None:
        anonWs = anonWb.get_sheet_by_name("Entries")
    originalDir = get_dir(ORIGINAL_FOLDER_PATH)
    patients = get_patients(originalDir)
    print("Found patients: " + str(patients))
    if anonWb is None or anonWs is None:
        print("[ERROR] Workbook or Worksheet for entries is None")
        return

    for i, f in enumerate(originalDir):
        patient_id = get_patient_id(anonWs, patients[i])
        if patient_id is None:
            patient_id = generate_UUID4()
            create_anon_entry(anonWb, anonWs, patients[i], patient_id)
        save_anon_patient_file(f, patient_id)

    anonWb.save(ANON_ENTRIES_PATH)
